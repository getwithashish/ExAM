# export_xlsx_service.py
import datetime
import uuid
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import NamedStyle
from user_auth.models import User
from asset.models.memory import Memory
from asset.models.business_unit import BusinessUnit
from asset.models.location import Location
from asset.models.employee import Employee
from asset.models import Asset

class ExportXLSX:
    @staticmethod
    def export_xlsx(assets, expiry_dates):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Assets"
    
        # Define field names and insert 'expiry_date' after 'warranty_period'
        fields = [field.name for field in Asset._meta.fields]
        warranty_idx = fields.index('warranty_period')
        fields.insert(warranty_idx + 1, 'expiry_date')
    
        # Create a named style for datetime formatting
        datetime_style = NamedStyle(
            name="datetime_style", number_format="YYYY-MM-DD HH:MM:SS"
        )
    
        # Apply the datetime style to the columns containing datetime values
        datetime_columns = [
            "created_at",
            "updated_at",
            "date_of_purchase",
        ]  # Replace with the actual names of datetime fields
        for col_idx, field in enumerate(fields, start=1):
            if field in datetime_columns:
                ws.column_dimensions[
                    openpyxl.utils.get_column_letter(col_idx)
                ].width = 25
                ws.cell(row=1, column=col_idx).style = datetime_style
    
        # Write headers
        for idx, field in enumerate(fields, start=1):
            ws.cell(row=1, column=idx).value = field
    
        # Write asset details and expiry dates
        for row_idx, (asset, expiry_date) in enumerate(zip(assets, expiry_dates), start=2):
            for col_idx, field in enumerate(fields, start=1):
                if field == 'expiry_date':
                    value = expiry_date
                else:
                    value = getattr(asset, field)
                    if isinstance(value, uuid.UUID):
                        value = str(value)  # Convert UUID to string
                    elif field == "asset_type":
                        value = str(value)  # Assuming `asset_type` is a string representation of AssetType
                    elif isinstance(value, (Employee, Location, BusinessUnit, Memory, User)):
                        value = str(value)  # Convert object to string representation
                    elif value == "" or value is None:
                        value = ""  # Set empty string for empty values
                    elif isinstance(value, datetime.datetime):
                        if field in datetime_columns:
                            value = value.strftime("%Y-%m-%d %H:%M:%S.%f")  # Use '%f' for microseconds
                        else:
                            value = value.strftime("%Y-%m-%d")  # Use '%Y-%m-%d' for other dates
                ws.cell(row=row_idx, column=col_idx).value = value
    
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="assets.xlsx"'
    
        wb.save(response)
    
        return response
